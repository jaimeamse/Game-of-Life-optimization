import json
from openpyxl import Workbook
with open("result.json", "r") as f:
    datos = json.load(f)

wb = Workbook()
ws = wb.active

# Escribir datos en celdas
ws["A1"] = "Nombre Experimento"
ws["B1"] = "X"
ws["C1"] = "Y"
ws["D1"] = "Iteraciones"
ws["E1"] = "Ciclos (G)"
ws["F1"] = "Instrucciones (G)"
ws["G1"] = "IPC"
ws["H1"] = "Tiempo"
ws["I1"] = "Reloj"
ws["J1"] = "(t)/(it*ciclo)"
ws["K1"] = "N_threads"
for i in datos.keys():
    partesLlave = i.split("x")
    fila = [datos[i]["nameExperiment"], partesLlave[0], partesLlave[1], datos[i]["iterations"], datos[i]["ciclos"], datos[i]["instructions"], datos[i]["IPC"], datos[i]["tiempo"], datos[i]["reloj"], (10**9)*(float(datos[i]["tiempo"])/(float(datos[i]["iterations"])*int(partesLlave[0])*int(partesLlave[1]))), datos[i]["n_threads"]]
    ws.append(fila)
    if datos[i]["nameExperiment"] == "King_Batch":
        ws.append([""])

# Guardar el archivo
wb.save("resultado.xlsx")
